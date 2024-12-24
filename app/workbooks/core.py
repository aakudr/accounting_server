from datetime import datetime
from fastapi import UploadFile
from openpyxl import load_workbook
from io import BytesIO

from openpyxl.worksheet.worksheet import Worksheet

from workbooks.model import SpendingEntry, SpendingTable
from workbooks.utils import getMergedCellVal


async def read_workbook_from_bytes(file: UploadFile):
    # Read the uploaded file into memory
    file_content = await file.read()

    # Load the workbook using openpyxl
    workbook = load_workbook(filename=BytesIO(file_content))

    # Process the first sheet
    sheet = workbook.active
    return sheet


async def read_workbook(filename: str):
    # Load the workbook using openpyxl
    workbook = load_workbook(filename=filename)

    # Process the first sheet
    sheet = workbook.active
    return sheet


def workbook_to_rows(sheet: Worksheet):
    if sheet is None:
        return []
    results = []
    for row in sheet.iter_rows():
        # Convert each row into a dictionary (or process as needed)
        results.append(row)
    return results


def build_spending_table_from_results(results, sheet: Worksheet):
    spending_table = SpendingTable()
    spending_table.entries = []

    header_row = results[0]
    header_row_2 = results[1]

    for row_i in range(len(results)):
        row = results[row_i]
        if (row_i == 0) or (row_i == 1):
            continue

        for col_i in range(len(results[row_i])):
            cell = row[col_i]

            if col_i == 0 or col_i == 1:
                continue

            cell_value = cell.value
            if not isinstance(cell_value, (int, float)):
                try:
                    cell_value = int(cell_value)
                except:
                    continue

            print(sheet)

            # entry.category = header_row_2[col_i].value or header_row[col_i].value
            level1category = getMergedCellVal(sheet, header_row[col_i]) or ""
            level2category = getMergedCellVal(sheet, header_row_2[col_i]) or ""
            aggregated_category = (
                f"{level1category} ({level2category})"
                if isinstance(level2category, str)
                and len(level2category) > 0
                and str(level2category) != str(level1category)
                else str(level1category)
            )

            price = cell.value

            entry = SpendingEntry(category=aggregated_category, price=price)

            if cell.comment:
                entry.comment = cell.comment.text

            if isinstance(row[0].value, datetime):
                entry.date = row[0].value
            else:
                continue

            if entry.price > 0 and entry.category and entry.date:
                print(entry.price, entry.category, entry.date)
                spending_table.entries.append(entry)

    print(len(spending_table.entries))
    return spending_table
